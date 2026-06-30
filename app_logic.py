"""
APPLICATION LAYER
Business logic: CSV/Excel parsing, trade reconstruction, tag definitions.
No HTTP, no SQL — only pure domain logic.
"""

import csv
import io
import json
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import database as db


# ── Tag Definitions ───────────────────────────────────────────────────────────

TAG_GROUPS = [
    {
        "id": "with",
        "label": "Technicals",
        "dot": "dot-with",
        "active_class": "active-with",
        "tags": ["Value", "Market Internals", "ADH", "AVWAP", "VWAP"],
        "multi": True,
    },
    {
        "id": "volume",
        "label": "Volume",
        "dot": "dot-vol",
        "active_class": "active-vol",
        "tags": ["Avg", "Above Avg", "Below Avg"],
        "multi": False,
    },
    {
        "id": "exit",
        "label": "Exit",
        "dot": "dot-exit",
        "active_class": "active-exit",
        "tags": ["Planned — Monitored Continuation", "Fear / Anxious"],
        "multi": False,
    },
    {
        "id": "setup",
        "label": "Setup",
        "dot": "dot-setup",
        "active_class": "active-setup",
        "tags": [
            "With Value", "Recapture of VWAP", "Break out of Range", "Initiative",
            "Low Tempo fade", "Balance Fade", "Look out of balance failed",
            "Gap fill failed", "No Setup", "Intuitive / Gut Feel"
        ],
        "multi": False,
    },
    {
        "id": "pre",
        "label": "Pre-Trade",
        "dot": "dot-pre",
        "active_class": "active-pre",
        "tags": [
            "Trade came to me", "Intuition / Mkt Feel", "Not sure about context",
            "Quick Profit Attitude", "Revenge Mindset", "Boredom", "Distracted"
        ],
        "multi": True,
    },
]


def get_tag_groups():
    """
    Return TAG_GROUPS with tags replaced by any custom DB config.
    Falls back to hardcoded defaults for groups not yet configured.
    """
    custom = db.get_tag_config()  # {group_id: [tag, ...]} or None
    if not custom:
        return TAG_GROUPS
    result = []
    for g in TAG_GROUPS:
        merged = dict(g)
        if g["id"] in custom:
            merged["tags"] = custom[g["id"]]
        result.append(merged)
    return result




OBSERVATION_CATEGORIES = [
    "price-action", "market-structure", "volume",
    "mkt internals", "psychology", "nuances", "general"
]

OBS_CATEGORY_GROUP = {
    "id": "obs_categories",
    "label": "Observation Categories",
    "tags": OBSERVATION_CATEGORIES,
    "multi": False,
    "dot": "dot-obs",
    "active_class": "active-obs",
}

OBSERVATION_GROUPS = ["Nuance", "Psychology", "Missed Opportunity", "Trade Setup"]

OBS_GROUP_GROUP = {
    "id": "obs_groups",
    "label": "Observation Groups",
    "tags": OBSERVATION_GROUPS,
    "multi": False,
    "dot": "dot-obs-grp",
    "active_class": "active-obs-grp",
}


def get_observation_categories():
    """Return observation categories, using custom DB config if available."""
    custom = db.get_tag_config()
    if custom and "obs_categories" in custom:
        return custom["obs_categories"]
    return OBSERVATION_CATEGORIES


def get_observation_groups():
    """Return observation groups, using custom DB config if available."""
    custom = db.get_tag_config()
    if custom and "obs_groups" in custom:
        return custom["obs_groups"]
    return OBSERVATION_GROUPS


# ── Day Marker Definitions ──────────────────────────────────────────────────

DAY_TYPE_TAGS = ["Balancing", "Trending", "Short Covering", "Liquidation", "Gap day", "Old business", "New Money"]
DAY_VALUE_TAGS = ["Lower", "OL", "Overlapping", "OH", "Higher"]
DAY_VOLUME_TAGS = ["Below Avg", "Avg", "Above Avg"]

DAY_TYPE_GROUP = {
    "id": "day_type",
    "label": "Day Type",
    "tags": DAY_TYPE_TAGS,
    "multi": True,
    "dot": "dot-day-type",
    "active_class": "active-vol",
}

DAY_VALUE_GROUP = {
    "id": "day_value",
    "label": "Value",
    "tags": DAY_VALUE_TAGS,
    "multi": False,
    "dot": "dot-day-value",
    "active_class": "active-setup",
}

DAY_VOLUME_GROUP = {
    "id": "day_volume",
    "label": "Volume",
    "tags": DAY_VOLUME_TAGS,
    "multi": False,
    "dot": "dot-day-vol",
    "active_class": "active-vol",
}


def get_day_type_tags():
    custom = db.get_tag_config()
    if custom and "day_type" in custom:
        return custom["day_type"]
    return DAY_TYPE_TAGS


def get_day_value_tags():
    custom = db.get_tag_config()
    if custom and "day_value" in custom:
        return custom["day_value"]
    return DAY_VALUE_TAGS


def get_day_volume_tags():
    custom = db.get_tag_config()
    if custom and "day_volume" in custom:
        return custom["day_volume"]
    return DAY_VOLUME_TAGS


DAY_GRADE_CATEGORIES = [
    {"name": "Market Read", "hint": "How well did you read market conditions?"},
    {"name": "Patience", "hint": "Waiting for setups, not forcing trades"},
    {"name": "Entry Quality", "hint": "Quality of trade entries"},
    {"name": "Risk Management", "hint": "Position sizing and stop placement"},
    {"name": "Exit Discipline", "hint": "Following exit rules"},
    {"name": "Emotional Control", "hint": "Managing emotions during trading"},
]

DAY_GRADE_GROUP = {
    "id": "grade_categories",
    "label": "Grade Categories",
    "tags": [c["name"] for c in DAY_GRADE_CATEGORIES],
    "multi": False,
    "dot": "dot-grade-cat",
    "active_class": "active-grade-cat",
}

GRADE_VALUES = {"Poor": 1, "Avg": 2, "Good": 3, "Excellent": 4}


def get_grade_categories():
    """Return list of grade category names (custom or default)."""
    custom = db.get_tag_config()
    if custom and "grade_categories" in custom:
        return custom["grade_categories"]
    return [c["name"] for c in DAY_GRADE_CATEGORIES]


def get_grade_categories_with_hints():
    """Return list of {name, hint} dicts for grade categories."""
    custom = db.get_tag_config()
    if custom and "grade_categories" in custom:
        default_hints = {c["name"]: c["hint"] for c in DAY_GRADE_CATEGORIES}
        return [{"name": n, "hint": default_hints.get(n, "")} for n in custom["grade_categories"]]
    return DAY_GRADE_CATEGORIES


DAY_CHECK_ITEMS = ["calm", "mkt_read", "awareness", "take_offer"]
DAY_CHECK_TOTAL = len(DAY_CHECK_ITEMS)  # 4


# ── Execution Score Helpers (v1/v2) ──────────────────────────────────────────

def get_execution_score_version(score_json):
    """Returns the version of an execution score JSON record."""
    return score_json.get("version", 1) if score_json else 1


def build_entry_execution_score(strength_record):
    """Builds v2 execution score JSON at trade entry from a strength record dict."""
    patience = strength_record.get("patience") or 0
    arrival = strength_record.get("arrival_context") or 0
    confirmation = strength_record.get("confirmation") or 0
    process_score = patience + arrival + confirmation

    return {
        "version": 2,
        "process": {
            "patience": bool(patience),
            "arrival_context": bool(arrival),
            "confirmation": bool(confirmation),
            "score": process_score
        },
        "technical": {
            "value": bool(strength_record.get("value") or 0),
            "volume": bool(strength_record.get("volume") or 0),
            "trend": bool(strength_record.get("trend") or 0),
            "adh": bool(strength_record.get("adh") or 0)
        },
        "mental_state": strength_record.get("mental_state") or "calm",
        "confidence": strength_record.get("confidence") or "medium",
        "management": None,
        "exit": None,
        "total_score": process_score,
        "max_score": 5
    }


def update_review_score(score_json, management_state, exit_quality):
    """Updates execution_score_json with post-trade review answers.
    Handles both v1 and v2 records. Returns the updated dict."""
    if not score_json:
        score_json = {}

    version = get_execution_score_version(score_json)

    management_score = 1 if management_state == "calm_objective" else 0
    exit_score = 1 if exit_quality == "planned" else 0

    if version == 2:
        score_json["management"] = {
            "state": management_state,
            "score": management_score
        }
        score_json["exit"] = {
            "quality": exit_quality,
            "score": exit_score
        }
        process_score = score_json.get("process", {}).get("score", 0)
        score_json["total_score"] = process_score + management_score + exit_score
        score_json["max_score"] = 5
    else:
        score_json["management"] = {
            "state": management_state,
            "score": management_score
        }
        score_json["exit"] = {
            "quality": exit_quality,
            "score": exit_score
        }
        score_json["total_score"] = management_score + exit_score
        score_json["max_score"] = 5
        score_json["_upgraded"] = True

    return score_json


def get_trade_execution_score(score_json):
    """Returns (actual_score, max_score) for a trade's execution score.
    Handles v1 and v2 formats gracefully."""
    if not score_json:
        return 0, 5

    version = get_execution_score_version(score_json)

    if version == 2:
        return score_json.get("total_score", 0), 5
    else:
        old_score = score_json.get("score", 0)
        return old_score, 5


def compute_day_score(scores_json):
    """Compute day checklist count from JSON.

    New format: {"calm":true,"mkt_read":false,...} → returns (checked_count, 4)
    Old format: {"Market Read":"Good",...} → returns None (graceful degrade)
    Returns (checked, total) tuple or None if no/invalid data.
    """
    import json
    if not scores_json:
        return None
    try:
        scores = json.loads(scores_json)
    except (json.JSONDecodeError, TypeError):
        return None
    if not isinstance(scores, dict):
        return None

    # Detect old format (values are strings like "Good", "Poor")
    for val in scores.values():
        if isinstance(val, str):
            return None  # old format, degrade gracefully

    checked = sum(1 for k in DAY_CHECK_ITEMS if scores.get(k) is True)
    return (checked, DAY_CHECK_TOTAL)


def compute_combined_day_score(day_score_json, trades):
    """Compute combined day score as a percentage.

    New formula (v2 trades present):
      Average execution score across trades as % of max.

    Legacy formula (all v1 trades):
      Process checklist (40%) + trade execution scores (60%).

    Mixed days (some v1, some v2): use new formula for all,
    treating v1 scores as-is out of 5.
    Returns integer percentage or None if nothing scored.
    """
    import json as _json

    if not trades:
        # Fall back to process-only if day checklist exists
        day_result = compute_day_score(day_score_json)
        if day_result is not None:
            checked, total = day_result
            return round((checked / total) * 100)
        return None

    # Check if any v2 trades exist
    has_v2 = False
    for t in trades:
        es_raw = t.get("execution_score_json") or t.get("exec_score_json")
        if es_raw:
            parsed = es_raw if isinstance(es_raw, dict) else None
            if parsed is None:
                try:
                    parsed = _json.loads(es_raw)
                except (ValueError, TypeError):
                    parsed = {}
            if get_execution_score_version(parsed) == 2:
                has_v2 = True
                break

    if has_v2:
        total_actual = 0
        total_max = 0
        for t in trades:
            es_raw = t.get("execution_score_json") or t.get("exec_score_json")
            parsed = None
            if es_raw:
                parsed = es_raw if isinstance(es_raw, dict) else None
                if parsed is None:
                    try:
                        parsed = _json.loads(es_raw)
                    except (ValueError, TypeError):
                        parsed = {}
            actual, mx = get_trade_execution_score(parsed)
            total_actual += actual
            total_max += mx
        return round((total_actual / total_max * 100)) if total_max > 0 else None
    else:
        # Legacy v1 formula — preserved exactly
        process_pct = None
        day_result = compute_day_score(day_score_json)
        if day_result is not None:
            checked, total = day_result
            process_pct = checked / total

        exec_pct = None
        exec_num = 0
        exec_den = 0
        for t in (trades or []):
            score = t.get("exec_score")
            if score is not None:
                exec_num += score
                exec_den += 5
        if exec_den > 0:
            exec_pct = exec_num / exec_den

        if process_pct is None and exec_pct is None:
            return None

        if exec_pct is not None and process_pct is not None:
            combined = exec_pct * 0.6 + process_pct * 0.4
        elif exec_pct is not None:
            combined = exec_pct
        else:
            combined = process_pct

        return round(combined * 100)


REQUIRED_COLUMNS = {"B/S", "avgPrice", "filledQty", "Fill Time", "Date"}


def parse_uploaded_file(filename: str, file_bytes: bytes) -> list:
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext == "csv":
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif ext in ("xlsx", "xls"):
        if not HAS_OPENPYXL:
            raise ValueError("openpyxl is required for Excel. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    else:
        raise ValueError(f"Unsupported file type: .{ext}. Upload CSV or XLSX.")

    if not rows:
        raise ValueError("File is empty.")

    missing = REQUIRED_COLUMNS - set(rows[0].keys())
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    fills = []
    for r in rows:
        if not r.get("Fill Time") or not r.get("B/S"):
            continue
        try:
            fills.append({
                "side":  str(r["B/S"]).strip(),
                "qty":   int(float(str(r["filledQty"]))),
                "price": float(r["avgPrice"]),
                "time":  _parse_fill_time(str(r["Fill Time"])),
                "date":  _parse_date(str(r["Date"])),
            })
        except (ValueError, KeyError, TypeError):
            continue

    if not fills:
        raise ValueError("No valid fills found in file.")

    return fills


def _parse_fill_time(raw: str) -> str:
    raw = raw.strip()
    for fmt in ("%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%m/%d/%y %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).strftime("%H:%M:%S")
        except ValueError:
            pass
    parts = raw.split()
    return parts[1] if len(parts) >= 2 else raw


def _parse_date(raw: str) -> str:
    raw = raw.strip().split()[0]
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return raw


# ── Trade Reconstruction ──────────────────────────────────────────────────────

def reconstruct_trades(fills: list) -> list:
    """Group fills by date → reconstruct round-trip trades per date."""
    by_date: dict = {}
    for f in fills:
        by_date.setdefault(f["date"], []).append(f)

    return [
        {"date": date, "trades": _build_round_trips(sorted(day_fills, key=lambda x: x["time"]))}
        for date, day_fills in sorted(by_date.items())
    ]


def _build_round_trips(fills: list) -> list:
    position = 0
    current  = []
    trades   = []

    for f in fills:
        position += f["qty"] if f["side"] == "Buy" else -f["qty"]
        current.append(f)
        if position == 0:
            trades.append(_compute_stats(current, len(trades) + 1))
            current = []

    if current:  # unclosed position
        trades.append(_compute_stats(current, len(trades) + 1))

    return trades


def _compute_stats(fills: list, trade_num: int) -> dict:
    buy_qty = buy_val = sell_qty = sell_val = 0
    for f in fills:
        if f["side"] == "Buy":
            buy_qty += f["qty"];  buy_val  += f["qty"] * f["price"]
        else:
            sell_qty += f["qty"]; sell_val += f["qty"] * f["price"]

    is_short  = fills[0]["side"] == "Sell"
    qty       = max(buy_qty, sell_qty)

    # Gracefully handle partial/unclosed positions (e.g. today's open trade)
    avg_entry = (sell_val / sell_qty) if (is_short and sell_qty) else (buy_val / buy_qty if buy_qty else 0)
    avg_exit  = (buy_val  / buy_qty)  if (is_short and buy_qty)  else (sell_val / sell_qty if sell_qty else 0)

    # P&L is 0 for unclosed positions (no exit side yet)
    if (is_short and buy_qty == 0) or (not is_short and sell_qty == 0):
        pnl = 0.0
    else:
        pnl = ((avg_entry - avg_exit) if is_short else (avg_exit - avg_entry)) * qty * 5

    return {
        "trade_num":  trade_num,
        "direction":  "Short" if is_short else "Long",
        "qty":        qty,
        "avg_entry":  round(avg_entry, 4),
        "avg_exit":   round(avg_exit,  4),
        "pnl":        round(pnl, 2),
        "entry_time": fills[0]["time"],
        "exit_time":  fills[-1]["time"],
        "fills":      fills,
        "open":       (is_short and buy_qty == 0) or (not is_short and sell_qty == 0),
    }


# ── Persistence ───────────────────────────────────────────────────────────────

def save_day_trades(date: str, trades: list, account_id=None) -> int:
    """
    Persist a full day of trades. Deletes and re-imports if day already exists
    for the same account.
    """
    existing = db.get_day_by_date_account(date, account_id)
    if existing:
        db.delete_day(existing["id"])

    day_id = db.upsert_day(date, account_id)

    for t in trades:
        trade_id = db.insert_trade(
            day_id, t["trade_num"], t["direction"], t["qty"],
            t["avg_entry"], t["avg_exit"], t["pnl"],
            t["entry_time"], t["exit_time"],
            is_open=t.get("open", False)
        )
        for f in t["fills"]:
            db.insert_fill(trade_id, f["time"], f["side"], f["qty"], f["price"])

    return day_id


def import_file(filename: str, file_bytes: bytes, account_id=None) -> dict:
    """Full pipeline: parse → reconstruct → save. Returns summary dict."""
    fills     = parse_uploaded_file(filename, file_bytes)
    day_trades = reconstruct_trades(fills)

    saved = []
    for d in day_trades:
        day_id = save_day_trades(d["date"], d["trades"], account_id)
        saved.append({
            "date":        d["date"],
            "day_id":      day_id,
            "trade_count": len(d["trades"]),
        })

    return {"days": saved}


# ── Live Trade Defaults & Config ─────────────────────────────────────────────

# Instrument tick values: (dollars_per_point, dollars_per_tick, ticks_per_point)
INSTRUMENT_CONFIG = {
    "MES": {"dollars_per_point": 5,  "dollars_per_tick": 1.25, "ticks_per_point": 4},
    "ES":  {"dollars_per_point": 50, "dollars_per_tick": 12.50, "ticks_per_point": 4},
}

# Default stop/TP distances in points
DEFAULT_TRADE_DEFAULTS = {
    "full_stop_points":    "20",
    "full_tp_points":      "20",
    "partial_stop_points": "20",
    "partial_tp1_points":  "5",
    "partial_tp2_points":  "10",
    "partial_tp3_points":  "20",
}


def get_trade_defaults():
    """Get trade default settings, merging DB config over hardcoded defaults."""
    config = db.get_all_config()
    result = {}
    for key, default_val in DEFAULT_TRADE_DEFAULTS.items():
        result[key] = config.get(f"td_{key}", default_val)
    return result


def get_instrument_config():
    """Get instrument tick values, with DB overrides."""
    config = db.get_all_config()
    result = {}
    for inst, defaults in INSTRUMENT_CONFIG.items():
        result[inst] = {
            "dollars_per_point": float(config.get(f"inst_{inst}_dpp", defaults["dollars_per_point"])),
            "dollars_per_tick":  float(config.get(f"inst_{inst}_dpt", defaults["dollars_per_tick"])),
            "ticks_per_point":   int(config.get(f"inst_{inst}_tpp",  defaults["ticks_per_point"])),
        }
    return result


# Default risk-stop distance (points) for the per-tranche risk capture.
DEFAULT_RISK_STOP_POINTS = 20.0


def compute_default_risk_stop(direction, entry_price, points=DEFAULT_RISK_STOP_POINTS):
    """Direction-aware default risk stop for an OPEN/ADD entry decision.
    Long → entry − points; Short → entry + points. Distinct from the working
    stop (live_trade_levels); used only for per-tranche risk capture."""
    entry_price = float(entry_price)
    is_long = str(direction).lower() == "long"
    stop = entry_price - points if is_long else entry_price + points
    return round(stop, 2)


def compute_tranche_risk(direction, instrument, exec_price, stop_price, entry_qty):
    """Risk for one OPEN/ADD row = |exec_price − stop_price| × entry_qty × $/point.
    Uses the row's full committed (entry) qty, not the current open qty. Returns
    None when there's no stop to derive from."""
    if stop_price is None:
        return None
    inst = get_instrument_config().get(instrument, INSTRUMENT_CONFIG["MES"])
    dpp = inst["dollars_per_point"]
    return round(abs(float(exec_price) - float(stop_price)) * int(entry_qty) * dpp, 2)


def compute_live_trade_plan(direction, instrument, entry_price, total_qty, mode):
    """
    Compute stop/TP levels with risk/reward for a new live trade.
    Returns list of level dicts ready for DB insertion.
    """
    inst = get_instrument_config().get(instrument, INSTRUMENT_CONFIG["MES"])
    dpp = inst["dollars_per_point"]
    defaults = get_trade_defaults()
    is_long = direction == "Long"
    levels = []

    if mode == "full":
        stop_dist = float(defaults["full_stop_points"])

        stop_price = entry_price - stop_dist if is_long else entry_price + stop_dist
        risk = abs(entry_price - stop_price) * total_qty * dpp

        levels.append({"level_type": "stop", "portion": 1, "qty": total_qty,
                        "price": round(stop_price, 2), "risk_dollars": round(risk, 2), "reward_dollars": 0})
    else:
        # Partials — divide into 3
        stop_dist = float(defaults["partial_stop_points"])
        tp1_dist  = float(defaults["partial_tp1_points"])
        tp2_dist  = float(defaults["partial_tp2_points"])
        tp3_dist  = float(defaults["partial_tp3_points"])

        # Divide qty as evenly as possible
        base_qty = total_qty // 3
        remainder = total_qty % 3
        qtys = [base_qty, base_qty, base_qty]
        # Distribute remainder to first portions
        for i in range(remainder):
            qtys[i] += 1

        tp_dists = [tp1_dist, tp2_dist, tp3_dist]

        for i in range(3):
            stop_price = entry_price - stop_dist if is_long else entry_price + stop_dist
            risk = abs(entry_price - stop_price) * qtys[i] * dpp

            levels.append({"level_type": "stop", "portion": i + 1, "qty": qtys[i],
                            "price": round(stop_price, 2), "risk_dollars": round(risk, 2), "reward_dollars": 0})

            tp_price = entry_price + tp_dists[i] if is_long else entry_price - tp_dists[i]
            reward = abs(tp_price - entry_price) * qtys[i] * dpp

            levels.append({"level_type": "tp", "portion": i + 1, "qty": qtys[i],
                            "price": round(tp_price, 2), "risk_dollars": 0, "reward_dollars": round(reward, 2)})

    return levels


def compute_execution_pnl(direction, instrument, entry_price, exec_price, qty):
    """Calculate P&L for a single execution."""
    inst = get_instrument_config().get(instrument, INSTRUMENT_CONFIG["MES"])
    dpp = inst["dollars_per_point"]
    if direction == "Long":
        return round((exec_price - entry_price) * qty * dpp, 2)
    else:
        return round((entry_price - exec_price) * qty * dpp, 2)


def recalculate_live_trade(live_trade):
    """
    Given a full live_trade dict (with levels + executions),
    compute remaining qty, current risk, potential profit, realized P&L.
    
    Risk calculation:
    - For each remaining portion, compute distance from entry to stop
    - If stop is on the LOSING side of entry → that's risk (negative, red)
    - If stop is on the WINNING side of entry (trailing stop past entry) → that's locked profit (positive, green)
    - Net risk = sum of all portion risks - total realized P&L already banked
    """
    total_qty   = live_trade["total_qty"]
    executions  = live_trade.get("executions", [])
    levels      = live_trade.get("levels", [])
    entry_price = live_trade["entry_price"]
    direction   = live_trade["direction"]
    instrument  = live_trade["instrument"]
    mode        = live_trade["mode"]

    inst = get_instrument_config().get(instrument, INSTRUMENT_CONFIG["MES"])
    dpp = inst["dollars_per_point"]
    is_long = direction == "Long"

    # Realized P&L and qty from executions
    exited_qty = sum(e["qty"] for e in executions)
    realized_pnl = sum(e["pnl"] for e in executions)
    remaining_qty = total_qty - exited_qty

    # Initial risk (what it was at entry, for reference)
    initial_risk = 0
    for lv in levels:
        if lv["level_type"] == "stop":
            dist = abs(entry_price - lv["price"]) * lv["qty"] * dpp
            initial_risk += dist

    if remaining_qty <= 0:
        return {
            "remaining_qty":   0,
            "exited_qty":      exited_qty,
            "realized_pnl":    round(realized_pnl, 2),
            "current_risk":    0,
            "potential_reward": 0,
            "initial_risk":    round(initial_risk, 2),
            "active_portions": [],
            "portion_details": [],
            "is_closed":       True,
        }

    # Figure out which portions are still open and their real risk
    portion_details = []  # detailed per-portion breakdown

    if mode == "partials":
        # Track how much has been exited per portion
        portion_exited = {1: 0, 2: 0, 3: 0}
        for e in executions:
            portion_exited[e["portion"]] = portion_exited.get(e["portion"], 0) + e["qty"]

        # Build lookup for levels
        portion_levels = {}
        for lv in levels:
            key = (lv["level_type"], lv["portion"])
            portion_levels[key] = lv

        total_stop_risk = 0  # raw risk from stops (can be negative = locked profit)
        total_reward = 0

        for p in [1, 2, 3]:
            stop_lv = portion_levels.get(("stop", p))
            tp_lv   = portion_levels.get(("tp", p))
            if not stop_lv:
                continue

            orig_qty = stop_lv["qty"]
            exited = portion_exited.get(p, 0)
            rem = orig_qty - exited
            if rem <= 0:
                continue

            # Calculate stop outcome: what happens if stopped out
            # For Long: stop below entry = loss, stop above entry = locked profit
            # For Short: stop above entry = loss, stop below entry = locked profit
            if is_long:
                stop_pnl = (stop_lv["price"] - entry_price) * rem * dpp
            else:
                stop_pnl = (entry_price - stop_lv["price"]) * rem * dpp

            # stop_pnl < 0 means risk (loss if stopped)
            # stop_pnl > 0 means locked profit (trailing stop past entry)
            total_stop_risk += stop_pnl

            # Reward from TP
            if tp_lv:
                if is_long:
                    tp_pnl = (tp_lv["price"] - entry_price) * rem * dpp
                else:
                    tp_pnl = (entry_price - tp_lv["price"]) * rem * dpp
                total_reward += max(tp_pnl, 0)
            else:
                tp_pnl = 0

            portion_details.append({
                "portion": p, "qty": rem,
                "stop_price": stop_lv["price"],
                "tp_price": tp_lv["price"] if tp_lv else None,
                "stop_pnl": round(stop_pnl, 2),  # negative=risk, positive=locked profit
                "tp_pnl": round(tp_pnl, 2),
            })

        # Net worst case = what happens if ALL remaining portions get stopped
        # plus what we've already realized
        worst_case_pnl = total_stop_risk + realized_pnl
        # Current risk = how much we could lose from HERE (negative = at risk, positive = net profitable even if stopped)
        # For display: current_risk is the absolute downside from current state
        if worst_case_pnl < 0:
            current_risk = abs(worst_case_pnl)
        else:
            current_risk = 0  # we're in profit even if stopped everywhere

        net_stop_exposure = round(worst_case_pnl, 2)

    else:
        # Full mode
        total_reward = 0
        net_stop_exposure = 0

        if remaining_qty > 0:
            stop_lv = next((lv for lv in levels if lv["level_type"] == "stop"), None)
            tp_lv   = next((lv for lv in levels if lv["level_type"] == "tp"), None)

            if stop_lv:
                if is_long:
                    stop_pnl = (stop_lv["price"] - entry_price) * remaining_qty * dpp
                else:
                    stop_pnl = (entry_price - stop_lv["price"]) * remaining_qty * dpp

                worst_case = stop_pnl + realized_pnl
                if worst_case < 0:
                    current_risk = abs(worst_case)
                else:
                    current_risk = 0
                net_stop_exposure = round(worst_case, 2)
            else:
                current_risk = 0

            if tp_lv:
                if is_long:
                    tp_pnl = (tp_lv["price"] - entry_price) * remaining_qty * dpp
                else:
                    tp_pnl = (entry_price - tp_lv["price"]) * remaining_qty * dpp
                total_reward = max(tp_pnl, 0)

        portion_details = []

    return {
        "remaining_qty":    remaining_qty,
        "exited_qty":       exited_qty,
        "realized_pnl":     round(realized_pnl, 2),
        "current_risk":     round(current_risk, 2),
        "potential_reward":  round(total_reward, 2),
        "initial_risk":     round(initial_risk, 2),
        "net_stop_exposure": net_stop_exposure,  # negative = net loss if all stopped, positive = net profit even if stopped
        "active_portions":  [p for p in portion_details if p["qty"] > 0] if mode == "partials" else [],
        "portion_details":  portion_details,
        "is_closed":        remaining_qty <= 0,
    }


def generate_shadow_trades(source_trade_id):
    """
    Generate shadow trades for all non-primary accounts based on a source trade.
    Source trade must belong to the primary account.
    """
    trade = db.get_trade_by_id(source_trade_id)
    if not trade:
        return

    primary = db.get_primary_account()
    if not primary:
        return

    # Verify source trade belongs to primary account
    if trade.get("account_id") != primary["id"]:
        return

    primary_qty = primary.get("default_qty") or trade["qty"]
    primary_instrument = primary.get("default_instrument") or "MES"
    inst_config = get_instrument_config()
    primary_dpp = inst_config.get(primary_instrument, inst_config["MES"])["dollars_per_point"]

    # Get price move in points
    if trade["direction"] == "Long":
        point_move = trade["avg_exit"] - trade["avg_entry"]
    else:
        point_move = trade["avg_entry"] - trade["avg_exit"]

    # Get all non-primary accounts with default_qty configured
    with db.get_conn() as conn:
        accounts = conn.execute(
            "SELECT * FROM accounts WHERE is_primary = 0 AND default_qty IS NOT NULL"
        ).fetchall()

    for p in accounts:
        p = dict(p)
        shadow_instrument = p.get("default_instrument") or "MES"
        shadow_dpp = inst_config.get(shadow_instrument, inst_config["MES"])["dollars_per_point"]
        shadow_default_qty = p["default_qty"]

        # Scale qty based on ratio of default qtys
        qty_ratio = shadow_default_qty / primary_qty if primary_qty else 1
        projected_qty = max(1, round(trade["qty"] * qty_ratio))

        # P&L = point_move × projected_qty × shadow_dpp
        projected_pnl = point_move * projected_qty * shadow_dpp

        db.upsert_shadow_trade(
            source_trade_id, p["id"],
            projected_qty, shadow_instrument, projected_pnl
        )


def regenerate_all_shadows():
    """Regenerate shadow trades for all existing trades in the primary account."""
    primary = db.get_primary_account()
    if not primary:
        return 0

    with db.get_conn() as conn:
        trades = conn.execute("""
            SELECT t.id FROM trades t
            JOIN trading_days d ON d.id = t.day_id
            WHERE d.account_id = ?
        """, (primary["id"],)).fetchall()

    count = 0
    for t in trades:
        generate_shadow_trades(t["id"])
        count += 1
    return count


def close_live_trade_to_journal(live_trade_id):
    """
    Explicitly save a live trade to the journal.
    Called by user clicking "Save & Push to Journal".
    Trade does NOT need to be fully closed — user decides when to push.
    """
    import json
    from datetime import date as dt_date

    lt = db.get_live_trade(live_trade_id)
    if not lt:
        return None

    calc = recalculate_live_trade(lt)

    # Use the trade's creation date (not today) so trades land on the correct day
    # created_at is stored as UTC; convert to local date via SQLite
    created_at = lt.get("created_at", "")
    if created_at:
        with db.get_conn() as conn:
            row = conn.execute(
                "SELECT date(?, 'localtime')", (created_at,)
            ).fetchone()
            trade_date = row[0] if row and row[0] else dt_date.today().isoformat()
    else:
        trade_date = dt_date.today().isoformat()
    account_id = lt.get("account_id")

    # Find or create trading day
    existing_day = db.get_day_by_date_account(trade_date, account_id)
    if existing_day:
        day_id = existing_day["id"]
    else:
        day_id = db.upsert_day(trade_date, account_id)

    # Determine next trade_num for this day
    with db.get_conn() as conn:
        max_num = conn.execute(
            "SELECT COALESCE(MAX(trade_num), 0) FROM trades WHERE day_id = ?", (day_id,)
        ).fetchone()[0]

    trade_num = max_num + 1

    # POC dynamic-trade-model: split executions into entry-side (OPEN/ADD) vs exit-side
    executions = lt.get("executions", [])
    entry_side_execs = [e for e in executions if str(e.get("exec_type") or "").upper() in ("OPEN", "ADD")]
    exit_side_execs = [e for e in executions if str(e.get("exec_type") or "").upper() not in ("OPEN", "ADD")]

    # Avg exit from exit-side executions only
    total_exit_val = sum(e["price"] * e["qty"] for e in exit_side_execs)
    total_exit_qty = sum(e["qty"] for e in exit_side_execs)
    avg_exit = round(total_exit_val / total_exit_qty, 4) if total_exit_qty else lt["entry_price"]

    # Get last exit execution time (fall back to entry_time if never exited)
    exit_time = exit_side_execs[-1]["exec_time"] if exit_side_execs else lt["entry_time"]

    realized_pnl = calc["realized_pnl"]

    # POC: prefer weighted_avg_entry (covers OPEN + ADDs); fall back to entry_price for legacy trades
    avg_entry_for_journal = lt.get("weighted_avg_entry") or lt["entry_price"]
    if not avg_entry_for_journal:
        avg_entry_for_journal = lt["entry_price"]

    # Build execution detail JSON for journal (levels + executions from live trade)
    levels = lt.get("levels", [])
    exec_detail = {
        "instrument": lt.get("instrument", "MES"),
        "mode": lt.get("mode", "full"),
        "entry_price": lt["entry_price"],
        "weighted_avg_entry": lt.get("weighted_avg_entry"),
        "levels": [{"level_type": lv["level_type"], "portion": lv["portion"],
                     "qty": lv["qty"], "price": lv["price"]} for lv in levels],
        "executions": [{"exec_type": e.get("exec_type",""), "portion": e["portion"],
                        "qty": e["qty"], "price": e["price"], "exec_time": e["exec_time"],
                        "pnl": e.get("pnl", 0)} for e in executions],
    }
    execution_json_str = json.dumps(exec_detail)

    trade_id = db.insert_trade(
        day_id, trade_num, lt["direction"], lt["total_qty"],
        avg_entry_for_journal, avg_exit, realized_pnl,
        lt["entry_time"], exit_time, is_open=(calc["remaining_qty"] > 0),
        execution_json=execution_json_str,
        execution_score_json=lt.get("execution_score_json"),
        context_id=lt.get("context_id")
    )

    # Save tags from live trade
    try:
        tags = json.loads(lt.get("tags_json", "{}"))
        for group_id, tag_list in tags.items():
            db.set_trade_tags(trade_id, group_id, tag_list)
    except (json.JSONDecodeError, AttributeError):
        pass

    # Save notes (all 3 fields)
    if lt.get("notes") or lt.get("notes_monitoring") or lt.get("notes_exit"):
        db.update_trade_notes(
            trade_id,
            lt.get("notes", ""),
            lt.get("notes_monitoring", ""),
            lt.get("notes_exit", ""),
        )

    # Generate fills from executions.
    # POC: when an OPEN row exists, all entry-side fills (OPEN + ADDs) come from the ledger.
    # Legacy trades (no OPEN row) still need a synthesised entry fill.
    entry_side = "Buy" if lt["direction"] == "Long" else "Sell"
    exit_side = "Sell" if lt["direction"] == "Long" else "Buy"

    if entry_side_execs:
        for e in entry_side_execs:
            db.insert_fill(trade_id, e["exec_time"], entry_side, e["qty"], e["price"], exit_type=None,
                           stop_price=e.get("stop_price"),
                           stop_source=e.get("stop_source") or "default")
    else:
        # Legacy path (no OPEN execution row): synthesise an entry fill. Carry the
        # live trade's initial working stop if one exists, else a 20-pt default.
        legacy_stops = [lv for lv in (lt.get("levels") or []) if lv.get("level_type") == "stop"]
        if legacy_stops:
            legacy_stop = legacy_stops[0].get("price")
            legacy_source = "default"
        else:
            legacy_stop = compute_default_risk_stop(lt["direction"], lt["entry_price"])
            legacy_source = "default"
        db.insert_fill(trade_id, lt["entry_time"], entry_side, lt["total_qty"], lt["entry_price"],
                       exit_type=None, stop_price=legacy_stop, stop_source=legacy_source)

    for e in exit_side_execs:
        db.insert_fill(trade_id, e["exec_time"], exit_side, e["qty"], e["price"],
                       exit_type=e.get("exec_type"), stop_price=None, stop_source='default')

    # Copy live trade images to journal trade
    live_images = db.get_live_trade_images(live_trade_id)
    for img in live_images:
        db.add_trade_image(trade_id, img['filename'], img['caption'])

    # Mark live trade as closed and link to journal
    db.update_live_trade(live_trade_id,
                         status="closed",
                         closed_at=dt_date.today().isoformat(),
                         realized_pnl=realized_pnl,
                         journal_trade_id=trade_id)

    # Generate shadow trades for non-primary accounts
    generate_shadow_trades(trade_id)

    return trade_id


# ══════════════════════════════════════════════════════════════════════════════
#  WEEKLY REVIEW (V1) — config + deterministic story engine
#  Pure Python, local-first, no model/network. Same input → same output.
# ══════════════════════════════════════════════════════════════════════════════

# Tunable config (defaults calibrated to the real data). Tag lists are editable in
# Settings via app_config; the numeric knobs live here so they can be tuned in code.
DEFAULT_IMPULSE_TAGS = ["Eager to trade", "Revenge Mindset", "Quick Profit Attitude"]
DEFAULT_OPERATIONAL_TAGS = ["Operational Error"]
OUTLIER_LOSS_MULT = 5      # single-trade loss vs trailing avg loss
MATERIAL_USD = 200         # buckets/insights below this absolute impact don't fire
STORY_MAX_SENTENCES = 4
WEEKLY_REVIEW_OBS_GROUP = "Review"   # extra allowed obs_group for collated review notes

# ── Trajectory tracking config + detector registry ───────────────────────────
# Qualifying-week floor: a week counts toward fired/not-fired only if it had ≥ this
# many trades. Configurable via app_config.
DEFAULT_QUALIFYING_FLOOR = 5
RECURRENCE_WINDOW_WEEKS = 4   # rolling (qualifying weeks)
TREND_WINDOW_WEEKS = 8        # rolling (qualifying weeks)
CHRONIC_PCT = 0.60           # fired in ≥60% of qualifying weeks across trend window
RESOLVED_SILENCE_WEEKS = 3   # qualifying weeks silent → a former leak counts as resolved
MIN_QUALIFYING_FOR_TREND = 4 # below this, classification is suppressed ("building")
IMPROVING_MIN_FIRINGS = 3    # need ≥ this many firings to call a magnitude trend


def _cfg_int(key, default, minimum=1):
    raw = db.get_config(key, "")
    try:
        v = int(raw)
        if v >= minimum:
            return v
    except (ValueError, TypeError):
        pass
    return default


def get_qualifying_floor():
    return _cfg_int("wr_qualifying_floor", DEFAULT_QUALIFYING_FLOOR)


def get_recurrence_window():
    return _cfg_int("wr_recurrence_window", RECURRENCE_WINDOW_WEEKS, minimum=2)


def get_trend_window():
    return _cfg_int("wr_trend_window", TREND_WINDOW_WEEKS, minimum=4)


def get_chronic_pct():
    raw = db.get_config("wr_chronic_pct", "")
    try:
        v = float(raw)
        if 0 < v <= 1:
            return v
    except (ValueError, TypeError):
        pass
    return CHRONIC_PCT


# Single source of truth: detector id → {label, polarity, tracked}. The trajectory
# zone iterates this; adding/removing a tracked pattern is a one-line edit here.
# net_result and concentration are weekly-only (not tracked).
DETECTOR_REGISTRY = {
    "net_result":        {"label": "Net result",            "polarity": "neutral",  "tracked": False},
    "operational_error": {"label": "Operational error",     "polarity": "leak",     "tracked": True},
    "impulsive_bucket":  {"label": "Impulsive trades",      "polarity": "leak",     "tracked": True},
    "revenge_chain":     {"label": "Revenge after loss",    "polarity": "leak",     "tracked": True},
    "oversized_loss":    {"label": "Oversized loss",        "polarity": "leak",     "tracked": True},
    "weak_exits":        {"label": "Weak exits",            "polarity": "leak",     "tracked": True},
    "expectancy_gap":    {"label": "Expectancy gap",        "polarity": "leak",     "tracked": True},
    "no_setup_leak":     {"label": "No-setup leak",         "polarity": "leak",     "tracked": True},
    "concentration":     {"label": "Concentration",         "polarity": "neutral",  "tracked": False},
    "came_to_me":        {"label": "Came to me",            "polarity": "strength", "tracked": True},
}


def tracked_detectors():
    """Detector ids that are tracked for trajectory, in registry order."""
    return [did for did, m in DETECTOR_REGISTRY.items() if m["tracked"]]


def get_impulse_tags():
    raw = db.get_config("wr_impulse_tags", "")
    if raw:
        try:
            val = json.loads(raw)
            if isinstance(val, list) and val:
                return val
        except (ValueError, TypeError):
            pass
    return list(DEFAULT_IMPULSE_TAGS)


def get_operational_tags():
    raw = db.get_config("wr_operational_tags", "")
    if raw:
        try:
            val = json.loads(raw)
            if isinstance(val, list) and val:
                return val
        except (ValueError, TypeError):
            pass
    return list(DEFAULT_OPERATIONAL_TAGS)


def _all_tags(trade):
    """Flatten a trade's {group_id: [tags]} into a flat set of tag strings."""
    out = set()
    for tags in (trade.get("tags") or {}).values():
        for t in tags:
            out.add(t)
    return out


def _money(v):
    """Signed dollar string, e.g. +$1,954 / -$8,000 (whole dollars)."""
    v = v or 0
    sign = "+" if v >= 0 else "-"
    return f"{sign}${abs(round(v)):,}"


def _money_abs(v):
    return f"${abs(round(v or 0)):,}"


def compute_weekly_summary(trades):
    """Reduce a week's trades into the buckets the story engine + zones consume.
    Operational-tagged trades are split out of all 'discretionary' stats."""
    impulse_tags = set(get_impulse_tags())
    operational_tags = set(get_operational_tags())

    for t in trades:
        t["_tagset"] = _all_tags(t)
        t["_operational"] = bool(t["_tagset"] & operational_tags)
        t["_impulse"] = bool(t["_tagset"] & impulse_tags)

    disc = [t for t in trades if not t["_operational"]]
    oper = [t for t in trades if t["_operational"]]

    def _net(ts):
        return round(sum(t["pnl"] for t in ts), 2)

    net_all = _net(trades)
    operational_pnl = _net(oper)
    discretionary_pnl = round(net_all - operational_pnl, 2)

    wins = [t for t in trades if t["pnl"] > 0]
    losses = [t for t in trades if t["pnl"] < 0]
    decided = len(wins) + len(losses)
    win_rate = round(100.0 * len(wins) / decided, 1) if decided else 0.0

    disc_losses = [t for t in disc if t["pnl"] < 0]
    avg_loss = round(sum(abs(t["pnl"]) for t in disc_losses) / len(disc_losses), 2) if disc_losses else 0.0
    disc_wins = [t for t in disc if t["pnl"] > 0]
    avg_win = round(sum(t["pnl"] for t in disc_wins) / len(disc_wins), 2) if disc_wins else 0.0

    days = sorted({t["date"] for t in trades})

    impulse = [t for t in disc if t["_impulse"]]
    came_to_me = [t for t in disc if "Trade came to me" in t["_tagset"]]
    no_setup = [t for t in disc if "No Setup" in (t.get("tags") or {}).get("setup", [])]

    # Exit discipline buckets (exit-group tags)
    planned = [t for t in disc if any("Planned" in x for x in (t.get("tags") or {}).get("exit", []))]
    fear_bail = [t for t in disc if any(("Fear" in x or "Bailed" in x) for x in (t.get("tags") or {}).get("exit", []))]

    # Execution scores (present on a subset)
    scored = [t for t in trades if t.get("exec_score") is not None]
    avg_exec_score = round(sum(t["exec_score"] for t in scored) / len(scored), 1) if scored else None

    # Revenge-after-loss chain: an impulse/revenge trade immediately after a same-day loss
    # bigger than the trailing avg loss.
    revenge_chain = None
    by_day = {}
    for t in trades:
        by_day.setdefault(t["date"], []).append(t)
    threshold = avg_loss if avg_loss > 0 else 0
    for day, ts in by_day.items():
        ts_sorted = sorted(ts, key=lambda x: x.get("trade_num", 0))
        for i in range(1, len(ts_sorted)):
            cur, prev = ts_sorted[i], ts_sorted[i - 1]
            is_revenge = bool(cur["_tagset"] & {"Revenge Mindset", "Eager to trade"})
            if is_revenge and prev["pnl"] < 0 and abs(prev["pnl"]) > threshold:
                if revenge_chain is None or abs(prev["pnl"]) > revenge_chain["prior_loss"]:
                    revenge_chain = {"prior_loss": abs(prev["pnl"]), "date": day,
                                     "trade_id": cur["id"], "prior_trade_id": prev["id"]}

    # Outlier loss: a single non-operational loss > OUTLIER_LOSS_MULT × trailing avg loss
    outlier = None
    if avg_loss > 0:
        for t in disc_losses:
            if abs(t["pnl"]) > OUTLIER_LOSS_MULT * avg_loss:
                if outlier is None or abs(t["pnl"]) > abs(outlier["pnl"]):
                    outlier = t

    # Sign-flip concentration on discretionary trades
    concentration = None
    if disc:
        biggest = max(disc, key=lambda t: abs(t["pnl"]))
        without = round(discretionary_pnl - biggest["pnl"], 2)
        if discretionary_pnl != 0 and (without >= 0) != (discretionary_pnl >= 0):
            concentration = {"without": without, "biggest_pnl": biggest["pnl"],
                         "swing": abs(biggest["pnl"])}

    return {
        "trades": trades,
        "discretionary": disc,
        "operational": oper,
        "net_all": net_all,
        "operational_pnl": operational_pnl,
        "discretionary_pnl": discretionary_pnl,
        "wins": len(wins), "losses": len(losses), "win_rate": win_rate,
        "trade_count": len(trades), "day_count": len(days), "days": days,
        "avg_loss": avg_loss, "avg_win": avg_win,
        "impulse": impulse, "impulse_net": _net(impulse),
        "came_to_me": came_to_me, "came_to_me_net": _net(came_to_me),
        "no_setup": no_setup, "no_setup_net": _net(no_setup),
        "planned": planned, "planned_net": _net(planned),
        "fear_bail": fear_bail, "fear_bail_net": _net(fear_bail),
        "avg_exec_score": avg_exec_score, "scored_count": len(scored),
        "revenge_chain": revenge_chain, "outlier": outlier, "concentration": concentration,
    }


# ── Detectors: each returns dict {key, fired, salience, sentence} or None ──────

def _det_net_result(s):
    wl = f"{s['wins']}W/{s['losses']}L"
    sentence = (f"You netted {_money(s['net_all'])} over {s['trade_count']} "
                f"trade{'s' if s['trade_count'] != 1 else ''} and {s['day_count']} "
                f"day{'s' if s['day_count'] != 1 else ''} — a {wl} week "
                f"({s['win_rate']:.0f}% win rate).")
    return {"key": "net_result", "fired": True, "salience": float("inf"), "sentence": sentence}


def _det_operational_error(s):
    if not s["operational"]:
        return None
    sentence = (f"{_money_abs(s['operational_pnl'])} of that was an operational error, "
                f"not a trading decision — your discretionary trading was "
                f"{_money(s['discretionary_pnl'])}.")
    return {"key": "operational_error", "fired": True, "salience": abs(s["operational_pnl"]),
            "sentence": sentence, "pinned": 2}


def _det_impulsive_bucket(s):
    imp = s["impulse"]
    if not imp:
        return None
    net = s["impulse_net"]
    if not (net < 0 or abs(net) >= MATERIAL_USD):
        return None
    base = (f"{len(imp)} impulsive trade{'s' if len(imp) != 1 else ''} netted {_money(net)}")
    if net < 0:
        without = round(s["discretionary_pnl"] - net, 2)
        base += f"; without them your discretionary week was {_money(without)}."
    else:
        base += "."
    return {"key": "impulsive_bucket", "fired": True, "salience": abs(net), "sentence": base}


def _det_revenge_chain(s):
    rc = s["revenge_chain"]
    if not rc:
        return None
    sentence = (f"After the {_money_abs(rc['prior_loss'])} loss you took a revenge trade "
                f"right after — the classic tilt pattern.")
    return {"key": "revenge_chain", "fired": True, "salience": rc["prior_loss"], "sentence": sentence}


def _det_oversized_loss(s):
    o = s["outlier"]
    if not o:
        return None
    sentence = (f"A single trade lost {_money_abs(o['pnl'])} — more than "
                f"{OUTLIER_LOSS_MULT}× your average loss, the kind of outlier that defines a week.")
    return {"key": "oversized_loss", "fired": True, "salience": abs(o["pnl"]), "sentence": sentence}


def _det_concentration(s):
    sf = s["concentration"]
    if not sf or sf["swing"] < MATERIAL_USD:
        return None
    sentence = (f"Without your single biggest trade, the discretionary week was "
                f"{_money(sf['without'])} — one position carried the result.")
    return {"key": "concentration", "fired": True, "salience": sf["swing"], "sentence": sentence}


def _det_expectancy_gap(s):
    # Win rate is "lying": ≥50% wins but discretionary net negative.
    if s["win_rate"] < 50 or s["discretionary_pnl"] >= 0:
        return None
    if abs(s["discretionary_pnl"]) < MATERIAL_USD:
        return None
    mult = round(s["avg_loss"] / s["avg_win"], 1) if s["avg_win"] else 0
    tail = f" — avg loss {_money_abs(s['avg_loss'])} is {mult}× avg win {_money_abs(s['avg_win'])}" if mult else ""
    sentence = (f"You won {s['win_rate']:.0f}% of trades but still lost money{tail}.")
    return {"key": "expectancy_gap", "fired": True, "salience": abs(s["discretionary_pnl"]), "sentence": sentence}


def _det_weak_exits(s):
    if not s["fear_bail"] or s["fear_bail_net"] >= 0 or abs(s["fear_bail_net"]) < MATERIAL_USD:
        return None
    sentence = (f"Fear/bail-out exits cost {_money(s['fear_bail_net'])} while planned exits ran "
                f"{_money(s['planned_net'])} — the exits were emotional.")
    return {"key": "weak_exits", "fired": True, "salience": abs(s["fear_bail_net"]), "sentence": sentence}


def _det_no_setup_leak(s):
    if not s["no_setup"] or s["no_setup_net"] >= 0 or abs(s["no_setup_net"]) < MATERIAL_USD:
        return None
    sentence = (f"Trades with no named setup leaked {_money(s['no_setup_net'])} — "
                f"the no-setup trades are a recurring drain.")
    return {"key": "no_setup_leak", "fired": True, "salience": abs(s["no_setup_net"]), "sentence": sentence}


def _det_came_to_me(s):
    if not s["came_to_me"] or s["discretionary_pnl"] <= 0 or s["came_to_me_net"] < MATERIAL_USD:
        return None
    sentence = (f"Your 'came to me' trades netted {_money(s['came_to_me_net'])} — "
                f"the patient, setup-driven entries paid.")
    return {"key": "came_to_me", "fired": True, "salience": abs(s["came_to_me_net"]), "sentence": sentence}


_STORY_DETECTORS = [
    _det_net_result, _det_operational_error, _det_impulsive_bucket, _det_revenge_chain,
    _det_oversized_loss, _det_concentration, _det_expectancy_gap, _det_weak_exits,
    _det_no_setup_leak, _det_came_to_me,
]


def generate_story(summary):
    """Run all detectors, lead with net, pin operational to slot 2, then sort the
    rest by salience (absolute dollar impact). Returns (sentences, fired_keys)."""
    fired = [d(summary) for d in _STORY_DETECTORS]
    fired = [f for f in fired if f]

    lead = next((f for f in fired if f["key"] == "net_result"), None)
    pinned = next((f for f in fired if f.get("pinned") == 2), None)
    rest = [f for f in fired if f is not lead and f is not pinned]
    rest.sort(key=lambda f: f["salience"], reverse=True)

    ordered = []
    if lead:
        ordered.append(lead)
    if pinned:
        ordered.append(pinned)
    ordered.extend(rest)
    ordered = ordered[:STORY_MAX_SENTENCES]

    return [f["sentence"] for f in ordered], [f["key"] for f in fired]


# ── Proposed intentions: fired detectors → candidate rules ────────────────────

_INTENTION_MAP = {
    "operational_error": "Confirm contract = MES before every entry.",
    "revenge_chain": "No new trade for 10 min after a hard loss.",
    "no_setup_leak": "No trade without a named setup.",
    "impulsive_bucket": "Name the trigger out loud before any impulsive entry.",
    "oversized_loss": "Predefine the max loss; cap single-trade risk.",
    "weak_exits": "Exit on the plan, not on fear — set the exit before entry.",
    "expectancy_gap": "Cut losers faster — your average loss dwarfs your average win.",
}


def propose_intentions(fired_keys):
    """Map fired detector keys → candidate rules (deduped, deterministic order).
    Each proposed rule is stamped with its target detector id (the proposing
    detector *is* the target) so linkage is free."""
    out, seen = [], set()
    for key in fired_keys:
        rule = _INTENTION_MAP.get(key)
        if rule and rule not in seen:
            seen.add(rule)
            out.append({"text": rule, "source": "proposed", "detector_key": key, "targets": key})
    return out


# ── Weekly logging hook: persist each tracked detector to insight_log ─────────
# Signed magnitude + count per tracked detector, read from the canonical summary
# buckets (never re-deriving the fired condition — that comes from the detectors).

def _signal_for(detector_id, summary):
    """Return (signed_magnitude, count) for one tracked detector from the summary."""
    s = summary
    if detector_id == "operational_error":
        return s["operational_pnl"], len(s["operational"])
    if detector_id == "impulsive_bucket":
        return s["impulse_net"], len(s["impulse"])
    if detector_id == "revenge_chain":
        rc = s.get("revenge_chain")
        return (-rc["prior_loss"], 1) if rc else (0.0, 0)
    if detector_id == "oversized_loss":
        o = s.get("outlier")
        return (o["pnl"], 1) if o else (0.0, 0)
    if detector_id == "weak_exits":
        return s["fear_bail_net"], len(s["fear_bail"])
    if detector_id == "expectancy_gap":
        return s["discretionary_pnl"], s["losses"]
    if detector_id == "no_setup_leak":
        return s["no_setup_net"], len(s["no_setup"])
    if detector_id == "came_to_me":
        return s["came_to_me_net"], len(s["came_to_me"])
    return 0.0, 0


def persist_insight_log(account_id, week_start, summary, fired_keys):
    """Idempotent: write one insight_log row per tracked detector for this week.
    `fired` is reused from the detectors' own output (fired_keys); magnitude/count
    come from the summary. qualifying = week met the trade floor. A zero-trade week
    is not logged (no signal)."""
    if summary["trade_count"] < 1:
        return 0
    qualifying = 1 if summary["trade_count"] >= get_qualifying_floor() else 0
    fired_set = set(fired_keys)
    n = 0
    for detector_id in tracked_detectors():
        magnitude, count = _signal_for(detector_id, summary)
        db.upsert_insight_log(
            account_id, week_start, detector_id,
            1 if detector_id in fired_set else 0,
            magnitude, count, qualifying,
        )
        n += 1
    return n


def log_week_insights(account_id, week_start):
    """Compute a week's summary + story and persist its tracked-detector results.
    Standalone entry point for backfill; build_weekly_review_data also calls
    persist_insight_log inline so navigating a week keeps the log truthful."""
    mon, _sun = week_bounds(week_start)
    trades = db.get_trades_in_range(account_id, *week_bounds(week_start))
    summary = compute_weekly_summary(trades)
    _story, fired = generate_story(summary)
    return persist_insight_log(account_id, mon, summary, fired)


# ── State classification (§4) — pure functions over insight_log ──────────────
# Windows are over QUALIFYING weeks only (≥ floor trades). Light/no-trade weeks are
# skipped, never counted as "didn't fire". Rolling, not calendar.

def _qualifying_series(account_id, detector_id, as_of_week):
    """Qualifying insight_log rows for one detector up to as_of_week, oldest-first."""
    rows = db.get_insight_history(account_id, detector_id, "2000-01-01")
    return [r for r in rows if r["week_start"] <= as_of_week and r["qualifying"]]


def _slope(values):
    """Least-squares slope sign helper over an evenly-spaced series."""
    n = len(values)
    if n < 2:
        return 0.0
    xs = list(range(n))
    mx = sum(xs) / n
    my = sum(values) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, values))
    den = sum((x - mx) ** 2 for x in xs)
    return (num / den) if den else 0.0


def _is_improving(firing_mags, polarity):
    """Magnitude trending the 'good' way across the last few firings.
    Leak → |magnitude| falling; strength → magnitude rising. Uses the last
    IMPROVING_MIN_FIRINGS firings; needs at least that many (avoids w-o-w noise)."""
    if len(firing_mags) < IMPROVING_MIN_FIRINGS:
        return False
    recent = firing_mags[-IMPROVING_MIN_FIRINGS:]
    if polarity == "strength":
        return _slope(recent) > 0
    return _slope([abs(m) for m in recent]) < 0


def classify_detector_state(account_id, detector_id, as_of_week):
    """Classify one detector as of a week. Returns a dict with state (or None when
    history is too thin / the pattern is dormant) plus the data the UI needs."""
    meta = DETECTOR_REGISTRY.get(detector_id, {"label": detector_id, "polarity": "leak"})
    series = _qualifying_series(account_id, detector_id, as_of_week)
    n_qual = len(series)

    trend = series[-get_trend_window():]
    recur = series[-get_recurrence_window():]
    fired_flags = [r["fired"] for r in trend]
    fired_now = bool(trend and trend[-1]["fired"])
    fired_trend = sum(fired_flags)
    fired_recur = sum(r["fired"] for r in recur)
    firing_mags = [r["magnitude"] for r in trend if r["fired"]]
    last3 = trend[-RESOLVED_SILENCE_WEEKS:]
    latest_firing = next((r["magnitude"] for r in reversed(series) if r["fired"]), None)

    out = {
        "detector_id": detector_id,
        "label": meta["label"],
        "polarity": meta["polarity"],
        "qualifying_weeks": n_qual,
        "weeks_fired": fired_trend,
        "fired_now": fired_now,
        "latest_magnitude": latest_firing,
        "series": [{"week_start": r["week_start"], "magnitude": r["magnitude"],
                    "fired": r["fired"], "count": r["count"]} for r in trend],
        "enough_history": n_qual >= MIN_QUALIFYING_FOR_TREND,
        "state": None,
    }
    if n_qual < MIN_QUALIFYING_FOR_TREND:
        return out  # building — don't label noise as trend

    chronic = (fired_trend / len(trend)) >= get_chronic_pct()
    # was active earlier then silent for the last N qualifying weeks
    resolved = (len(trend) >= MIN_QUALIFYING_FOR_TREND
                and len(last3) == RESOLVED_SILENCE_WEEKS
                and not any(r["fired"] for r in last3)
                and fired_trend >= 2)
    improving = fired_now and _is_improving(firing_mags, meta["polarity"])
    recurring = fired_now and fired_recur >= 2
    new = fired_now and fired_recur == 1

    # Precedence: Resolved > Improving > Chronic > Recurring > New
    if resolved:
        out["state"] = "Resolved"
    elif improving:
        out["state"] = "Improving"
    elif chronic:
        out["state"] = "Chronic"
    elif recurring:
        out["state"] = "Recurring"
    elif new:
        out["state"] = "New"
    return out


def classify_all(account_id, as_of_week):
    """Classify every tracked detector as of a week. Returns {detector_id: result}."""
    return {did: classify_detector_state(account_id, did, as_of_week)
            for did in tracked_detectors()}


# ── Intention linkage (§5) — is the rule working? ────────────────────────────

def intention_linkage(account_id, targets, since_week, as_of_week):
    """Summarize the targeted detector's trajectory since an intention was set.
    `targets` is a detector id, or ''/'general' for manual-grade-only intentions.
    Evaluates qualifying weeks strictly AFTER since_week (the week the rule was set)."""
    if not targets or targets in ("general",) or targets not in DETECTOR_REGISTRY:
        return {"targets": targets or "", "has_target": False, "verdict": "no_target",
                "summary": "Self-graded only — no detector linked, so there's no automatic verdict.",
                "post_weeks": 0, "fired_after": 0}

    polarity = DETECTOR_REGISTRY[targets]["polarity"]
    series = _qualifying_series(account_id, targets, as_of_week)
    post = [r for r in series if r["week_start"] > since_week]
    fired_post = [r for r in post if r["fired"]]
    label = DETECTOR_REGISTRY[targets]["label"]

    base = {"targets": targets, "has_target": True, "label": label, "polarity": polarity,
            "post_weeks": len(post), "fired_after": len(fired_post),
            "series": [{"week_start": r["week_start"], "magnitude": r["magnitude"],
                        "fired": r["fired"]} for r in post]}

    if len(post) == 0:
        base["verdict"] = "too_soon"
        base["summary"] = "Set recently — not enough qualifying weeks yet to verify."
        return base

    if polarity == "strength":
        # working = the good behavior keeps showing up (and ideally growing)
        if fired_post:
            mags = [r["magnitude"] for r in fired_post]
            growing = _slope(mags) > 0 if len(mags) >= 2 else True
            base["verdict"] = "working" if growing else "mixed"
            base["summary"] = (f"'{label}' showed up in {len(fired_post)} of {len(post)} weeks since"
                               + (" and is growing." if growing else ", but isn't growing."))
        else:
            base["verdict"] = "not_working"
            base["summary"] = f"'{label}' hasn't shown up in the {len(post)} weeks since — the habit has slipped."
        return base

    # leak polarity
    if not fired_post:
        base["verdict"] = "working"
        base["summary"] = f"'{label}' hasn't recurred in the {len(post)} qualifying weeks since you set this. ✓"
    else:
        mags = [abs(r["magnitude"]) for r in fired_post]
        shrinking = _slope(mags) < 0 if len(mags) >= 2 else False
        if shrinking:
            base["verdict"] = "mixed"
            base["summary"] = (f"'{label}' still fired {len(fired_post)}× since, but the damage is shrinking.")
        else:
            base["verdict"] = "not_working"
            base["summary"] = (f"'{label}' fired {len(fired_post)} of {len(post)} weeks since — still leaking.")
    return base


# ── Trajectory zone builder (§6) ─────────────────────────────────────────────

def latest_trading_week(account_id):
    """Monday of the account's most recent trading day, or None."""
    d = db.get_latest_trade_date(account_id)
    if not d:
        return None
    return week_bounds(d)[0]


def _account_trading_mondays(account_id, up_to_week):
    """Distinct Mondays (ISO) the account traded on, ≤ up_to_week, oldest-first."""
    trades = db.get_trades_in_range(account_id, "2000-01-01", up_to_week)
    mondays = {week_bounds(t["date"])[0] for t in trades}
    return sorted(m for m in mondays if m <= up_to_week)


def backfill_insight_log(account_id, up_to_week=None, max_weeks=None):
    """Explicit, idempotent one-time backfill: log the account's trading weeks up to
    up_to_week (default = its latest trading week), so the trajectory has history from
    existing data. `max_weeks=None` backfills the full history; pass an int to limit to
    the most recent N weeks. Recompute is deterministic, so this derives history rather
    than inventing it. Re-running overwrites the same rows (UNIQUE upsert)."""
    if up_to_week is None:
        up_to_week = latest_trading_week(account_id)
        if up_to_week is None:
            return 0
    mondays = _account_trading_mondays(account_id, up_to_week)
    if max_weeks:
        mondays = mondays[-max_weeks:]
    for mon in mondays:
        log_week_insights(account_id, mon)
    return len(mondays)


# Display badge colors per LEAK state (CSS class suffixes).
STATE_BADGE = {
    "New": "cyan", "Recurring": "amber", "Chronic": "red",
    "Improving": "mint", "Resolved": "green",
}

# Strength patterns live in their own section and use their own display vocabulary
# (Building / Holding / Slipping) — never the leak words. This is display-only; the
# underlying computed state and the stored detector_id are unchanged.
STRENGTH_STATE = {
    "Improving": ("Building", "mint",  "strength building"),
    "New":       ("Building", "cyan",  "strength building"),
    "Chronic":   ("Holding",  "green", "holding steady"),
    "Recurring": ("Holding",  "green", "holding steady"),
    "Resolved":  ("Slipping", "amber", "good habit slipping"),
}


def _traj_row(result):
    """Shape one classification result into a trajectory display row. Leaks fall into
    the repeating / improving buckets; strengths get their own bucket with a Building /
    Holding / Slipping label so a declining strength reads as slipping, not 'repeating'."""
    state = result["state"]
    if state is None:
        return None
    pol = result["polarity"]
    mags = [p["magnitude"] for p in result["series"]]
    wf = result["weeks_fired"]
    # Frequency = qualifying weeks the pattern fired WITHIN the trailing window. The
    # denominator is the window size (= number of qualifying weeks of history seen,
    # capped at the trend window). Worded "X of N wks" so it can't read as weeks-ago.
    window = len(result["series"]) or 1
    freq = f"fired {wf} of {window} wks"
    if pol == "strength":
        disp = STRENGTH_STATE.get(state)
        if not disp:
            return None
        disp_state, badge, word = disp
        bucket, subtext = "strengths", f"{word} · {freq}"
    else:
        # Leak: Chronic/Recurring = repeating; Improving/Resolved = improving/beaten;
        # New leaks stay queryable but out of the two visible buckets (per brief).
        if state in ("Chronic", "Recurring"):
            bucket = "repeating"
        elif state in ("Improving", "Resolved"):
            bucket = "improving"
        else:
            return None
        disp_state, badge, subtext = state, STATE_BADGE[state], freq
    return {
        "detector_id": result["detector_id"], "label": result["label"],
        "polarity": pol, "state": state, "display_state": disp_state, "badge": badge,
        "subtext": subtext, "slipping": (pol == "strength" and state == "Resolved"),
        "weeks_fired": wf, "latest_magnitude": result["latest_magnitude"],
        "sparkline": mags, "series": result["series"], "bucket": bucket,
    }


def build_trajectory(account_id, anchor_week):
    """Assemble the trajectory zone anchored at anchor_week (most recent week).
    Two visible buckets capped at 2 each; plus intention-linkage cards."""
    classified = classify_all(account_id, anchor_week)

    # Gating: count qualifying weeks of history (max across detectors == weeks seen).
    qual_weeks = max((c["qualifying_weeks"] for c in classified.values()), default=0)
    needed = MIN_QUALIFYING_FOR_TREND
    if qual_weeks < needed:
        return {"visible": False, "qualifying_weeks": qual_weeks, "needed": needed,
                "anchor_week": anchor_week, "repeating": [], "improving": [],
                "strengths": [], "intentions": []}

    rows = [r for r in (_traj_row(c) for c in classified.values()) if r]
    repeating = sorted([r for r in rows if r["bucket"] == "repeating"],
                       key=lambda r: abs(r["latest_magnitude"] or 0), reverse=True)[:2]
    # Improving/beaten: Resolved (fully beaten) first, then by magnitude of what was tamed.
    improving = sorted([r for r in rows if r["bucket"] == "improving"],
                       key=lambda r: (r["state"] == "Resolved", abs(r["latest_magnitude"] or 0)),
                       reverse=True)[:2]
    # Strengths get their own section: slipping (caution) first, then by magnitude.
    strengths = sorted([r for r in rows if r["bucket"] == "strengths"],
                       key=lambda r: (r["slipping"], abs(r["latest_magnitude"] or 0)),
                       reverse=True)[:2]

    # Intention → pattern cards: intentions set in the trend window that target a
    # tracked detector, evaluated as of the anchor week.
    from datetime import date as _date, timedelta
    y, m, d = (int(x) for x in anchor_week.split("-"))
    since = (_date(y, m, d) - timedelta(weeks=get_trend_window())).isoformat()
    cards = []
    for it in db.get_intentions_in_range(account_id, since, anchor_week):
        link = intention_linkage(account_id, it.get("targets") or "", it["review_week"], anchor_week)
        cards.append({"id": it["id"], "text": it["text"], "source": it["source"],
                      "result": it["result"], "set_week": it["review_week"], **link})

    return {"visible": True, "qualifying_weeks": qual_weeks, "needed": needed,
            "anchor_week": anchor_week, "repeating": repeating, "improving": improving,
            "strengths": strengths, "intentions": cards}


# ── Weekly review data assembler (called by the route; thin SQL via db) ───────

def week_bounds(week_start):
    """Given a Monday ISO date string, return (monday, sunday) ISO strings."""
    from datetime import date as _date, timedelta
    y, m, d = (int(x) for x in week_start.split("-"))
    mon = _date(y, m, d)
    mon = mon - timedelta(days=mon.weekday())   # snap to Monday defensively
    sun = mon + timedelta(days=6)
    return mon.isoformat(), sun.isoformat()


def current_week_monday():
    from datetime import date as _date, timedelta
    today = _date.today()
    return (today - timedelta(days=today.weekday())).isoformat()


def _shift_week(week_start, weeks):
    from datetime import date as _date, timedelta
    y, m, d = (int(x) for x in week_start.split("-"))
    return (_date(y, m, d) + timedelta(weeks=weeks)).isoformat()


def _setup_table(disc_trades):
    by = {}
    for t in disc_trades:
        for setup in (t.get("tags") or {}).get("setup", []) or ["(none)"]:
            b = by.setdefault(setup, {"setup": setup, "count": 0, "net": 0.0, "wins": 0})
            b["count"] += 1
            b["net"] = round(b["net"] + t["pnl"], 2)
            if t["pnl"] > 0:
                b["wins"] += 1
    rows = list(by.values())
    for b in rows:
        b["win_rate"] = round(100.0 * b["wins"] / b["count"], 0) if b["count"] else 0
    rows.sort(key=lambda b: b["net"], reverse=True)
    return rows


def build_weekly_review_data(account_id, week_start):
    """Assemble the full weekly-review payload (KPIs, story, behavior, ledger,
    observations, themes, intentions). Pure read + deterministic story."""
    mon, sun = week_bounds(week_start)
    trades = db.get_trades_in_range(account_id, mon, sun)
    summary = compute_weekly_summary(trades)
    story, fired = generate_story(summary)

    # Hook: persist ONLY the week being viewed (idempotent upsert), so the log stays
    # truthful as weeks are viewed/recomputed. Historical backfill is a separate,
    # explicit step (run backfill_insights.py) — never triggered by viewing a page.
    persist_insight_log(account_id, mon, summary, fired)

    # Trajectory zone is a cross-week, trailing-window view anchored at the most
    # recent week — independent of the viewed week. Only surface it when viewing
    # that most-recent week, so a cross-week band never appears to belong to a past
    # week. It reads whatever is already in insight_log; until the hook has logged
    # ≥4 qualifying weeks (or you run the backfill), it shows the "building" state.
    latest_week = latest_trading_week(account_id) or mon
    is_current = mon >= latest_week
    if is_current:
        trajectory = build_trajectory(account_id, latest_week)
    else:
        trajectory = {"visible": False, "is_current": False,
                      "anchor_week": latest_week, "repeating": [], "improving": [],
                      "intentions": [], "qualifying_weeks": 0, "needed": MIN_QUALIFYING_FOR_TREND}
    trajectory["is_current"] = is_current

    impulse_tags = set(get_impulse_tags())
    operational_tags = set(get_operational_tags())

    # Ledger rows
    ledger = []
    for t in trades:
        tagset = t.get("_tagset") or _all_tags(t)
        pre = (t.get("tags") or {}).get("pre", [])
        setup = (t.get("tags") or {}).get("setup", [])
        ledger.append({
            "id": t["id"], "date": t["date"], "direction": t["direction"],
            "qty": t["qty"], "pnl": t["pnl"],
            "read": ", ".join(pre) if pre else "—",
            "setup": ", ".join(setup) if setup else "—",
            "operational": bool(tagset & operational_tags),
            "revenge": bool(tagset & {"Revenge Mindset", "Eager to trade"}),
            "exec_score": t.get("exec_score"),
        })

    # Observations in the week + review-group collation + theme counts
    observations = db.get_observations_in_range(mon, sun)
    review_notes = [o for o in observations if WEEKLY_REVIEW_OBS_GROUP in (o.get("group_list") or [])]

    # Daily reflection notes that exist (rare, but collate when present)
    day_reflections = []
    for day in db.get_all_days(mon, sun, account_id):
        full = db.get_day_by_id(day["id"]) or {}
        for fld, label in (("notes_well", "What went well"), ("notes_improve", "To improve"),
                           ("notes_lessons", "Lessons"), ("notes_focus", "Focus")):
            val = (full.get(fld) or "").strip()
            if val:
                day_reflections.append({"date": day["date"], "label": label, "text": val})

    theme_week = db.get_theme_counts(account_id, mon, sun)
    month_start = mon[:8] + "01"
    theme_month = db.get_theme_counts(account_id, month_start, sun)

    # Intentions: this week + last week (for grading the loop)
    review = db.get_or_create_weekly_review(account_id, mon)
    this_intentions = db.get_weekly_intentions(review["id"])
    last_mon = _shift_week(mon, -1)
    last_review = db.get_or_create_weekly_review(account_id, last_mon)
    last_intentions = db.get_weekly_intentions(last_review["id"])

    from datetime import date as _date
    y, m, d = (int(x) for x in mon.split("-"))
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    week_label = f"Week of {months[m-1]} {d}, {y}"

    return {
        "account_id": account_id,
        "week_start": mon, "week_end": sun,
        "prev_week": _shift_week(mon, -1), "next_week": _shift_week(mon, 1),
        "week_label": week_label,
        "kpis": {
            "net_all": summary["net_all"],
            "discretionary": summary["discretionary_pnl"],
            "operational": summary["operational_pnl"],
            "win_rate": summary["win_rate"],
            "wins": summary["wins"], "losses": summary["losses"],
            "trades": summary["trade_count"], "days": summary["day_count"],
        },
        "story": story,
        "behavior": {
            "came_to_me": {"count": len(summary["came_to_me"]), "net": summary["came_to_me_net"]},
            "impulse": {"count": len(summary["impulse"]), "net": summary["impulse_net"]},
            "avg_exec_score": summary["avg_exec_score"], "scored_count": summary["scored_count"],
            "planned_net": summary["planned_net"], "fear_bail_net": summary["fear_bail_net"],
            "setup_table": _setup_table(summary["discretionary"]),
        },
        "ledger": ledger,
        "observations": observations,
        "review_notes": review_notes,
        "day_reflections": day_reflections,
        "theme_week": theme_week, "theme_month": theme_month,
        "review_id": review["id"], "reflection_text": review["reflection_text"],
        "intentions": this_intentions,
        "last_week": {"week_start": last_mon, "review_id": last_review["id"],
                      "intentions": last_intentions},
        "proposed": propose_intentions(fired),
        "fired": fired,
        "trajectory": trajectory,
        "tracked_detectors": [{"id": did, "label": DETECTOR_REGISTRY[did]["label"]}
                              for did in tracked_detectors()],
    }
